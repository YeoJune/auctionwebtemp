import sys
import os
import requests
from requests.exceptions import RequestException
from bs4 import BeautifulSoup
import time
import urllib.parse
from PyQt5.QtWidgets import (QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout, QPushButton,
QListWidget, QLabel, QMessageBox, QProgressBar, QTableWidget, QTableWidgetItem,
QTabWidget, QFileDialog, QHeaderView)
from PyQt5.QtGui import QPixmap
from PyQt5.QtCore import Qt, QThread, pyqtSignal, QSize
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
from googletrans import Translator

# 브랜드 리스트 정의
BRANDS = [
    "LOUIS VUITTON", "CHANEL", "HERMES", "GUCCI", "PRADA", "ROLEX", "OMEGA", "TAG Heuer", "Cartier", "BVLGARI",
    "Tiffany＆Co.", "HARRY WINSTON", "Van Cleef＆Arpels", "A. LANGE & SOHNE", "AUDEMARS PIGUET", "Baccarat",
    "BALENCIAGA", "BALLY", "BAUME＆MERCIER", "Bell & Ross", "Berluti", "BOTTEGA VENETA", "BOUCHERON", "BREGUET",
    "BREITLING", "BUCCELLATI", "BURBERRY", "Carlo Parlati", "Carrera y Carrera", "CASIO", "CAZZANIGA", "CELINE",
    "CHARRIOL", "CHAUMET", "chloe", "Chopard", "Christian Louboutin", "CHROME HEARTS", "CITIZEN", "COACH",
    "Cole Haan", "COMME des GARCONS", "CORUM", "D＆G", "Damiani", "DE BEERS", "Dior", "DOLCE＆GABBANA", "DSQUARED2",
    "dunhill", "EDOX", "EMILIO PUCCI", "ETRO", "FEDERICO BUCCELLATI", "Felisi", "FENDI", "FRANCK MULLER", "FRED",
    "FREDERIQUE CONSTANT", "FURLA", "GaGa MILANO", "Georg Jensen", "gimel", "GIRARD PERREGAUX", "GIVENCHY",
    "GOYARD", "GRAFF", "GRAHAM", "HUBLOT", "IWC", "Jacob&co", "JAEGER LECOULTRE", "Jeunet", "JEWEL STUDIO",
    "JILSANDER", "JIMMY CHOO", "Justin Davis", "Kashikey", "Kate Spade", "LANVIN", "LOEWE", "LONG CHAMP",
    "LONGINES", "MACKINTOSH", "MARC BY MARC JACOBS", "MARC JACOBS", "MAUBOUSSIN", "MAURICE LACROIX", "MCM",
    "Meissen", "Michael Kors", "MIKIMOTO", "miu miu", "MONCLER", "MONTBLANC", "ORIENT", "Orobianco", "PANERAI",
    "PATEK PHILIPPE", "PIAGET", "POLA", "POMELLATO", "Ponte Vecchio", "RADO", "RayBan", "REGAL", "RICHARD MILLE",
    "RIMOWA", "Ritmo latino", "ROGER DUBUIS", "SAINT LAURENT", "Salvatore Ferragamo", "SEE BY CHLOE", "SEIKO",
    "Sergio Rossi", "SINN", "SOUTHERN CROSS", "Supreme", "TASAKI", "TOD'S", "Tom Ford", "Tory Burch", "TUDOR",
    "TUMI", "ULYSSE NARDIN", "UNIVERSAL GENEVE", "UNOAERRE", "VACHERON CONSTANTIN", "VALENTINO", "Vendome Aoyama",
    "Verite", "VERSACE", "WALTHAM", "Yves Saint Laurent", "ZENITH", "梶 光夫", "梶武史", "石川 暢子", "田村 俊一"]

def login_and_check():
    session = requests.Session()

    login_url = "https://www.ecoauc.com/client/users/post-sign-in"
    login_data = {
        "_method": "POST",
        "_csrfToken": "",
        "email_address": "casaboutique@naver.com",
        "password": "Wm8AXpVY2666",
        "remember-me": "remember-me"
    }

    headers = {
        "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/128.0.0.0 Safari/537.36",
        # ... (나머지 헤더 생략)
    }

    print("로그인 페이지 접근 중...")
    login_page = session.get("https://www.ecoauc.com/client/users/sign-in", headers=headers)
    print(f"로그인 페이지 상태 코드: {login_page.status_code}")

    soup = BeautifulSoup(login_page.text, 'html.parser')
    csrf_token = soup.find('input', {'name': '_csrfToken'})['value']
    login_data['_csrfToken'] = csrf_token
    print(f"CSRF 토큰 찾음: {csrf_token}")

    print("로그인 시도 중...")
    login_response = session.post(login_url, data=login_data, headers=headers, allow_redirects=False)
    print(f"Login status code: {login_response.status_code}")

    if login_response.status_code == 302:
        redirect_url = login_response.headers['Location']
        print(f"리다이렉트 URL: {redirect_url}")
        home_page = session.get(redirect_url, headers=headers)
        print(f"홈페이지 상태 코드: {home_page.status_code}")

        print("アカウント 페이지 접근 중...")
        account_url = "https://www.ecoauc.com/client/users"
        account_page = session.get(account_url, headers=headers)
        print(f"アカウント 페이지 상태 코드: {account_page.status_code}")

        soup = BeautifulSoup(account_page.text, 'html.parser')
        if "アカウント" in account_page.text and soup.find('a', string='ログアウト'):
            print("로그인 성공!")
            return session
        else:
            print("로그인 실패 또는 アカウント 페이지를 찾을 수 없습니다.")
            return None
    else:
        print("로그인 실패.")
        return None

def get_total_pages(session, base_url, params):
    response = session.get(base_url, params=params)
    soup = BeautifulSoup(response.text, 'html.parser')

    pagination = soup.find("ul", class_="pagination")
    if pagination:
        page_links = pagination.find_all("li")
        for link in reversed(page_links):
            a_tag = link.find("a")
            if a_tag and a_tag.text.strip().isdigit():
                return int(a_tag.text.strip())
    
    items = soup.find_all("div", class_="card")
    if items:
        return 1
    
    return 0

def create_output_folder(brand):
    output_folder = os.path.join(os.path.expanduser("~"), "Desktop", "크롤링결과", brand)
    os.makedirs(output_folder, exist_ok=True)
    return output_folder

def download_image(image_url, save_dir):
    try:
        parsed_url = urllib.parse.urlparse(image_url)
        query_params = urllib.parse.parse_qs(parsed_url.query)
        query_params.pop('w', None)
        query_params.pop('h', None)
        new_query = urllib.parse.urlencode(query_params, doseq=True)
        new_url = urllib.parse.urlunparse(parsed_url._replace(query=new_query))

        response = requests.get(new_url, timeout=30)
        if response.status_code == 200:
            os.makedirs(save_dir, exist_ok=True)
            file_name = os.path.join(save_dir, os.path.basename(new_url))
            with open(file_name, 'wb') as f:
                f.write(response.content)
            return file_name
    except Exception as e:
        print(f"Error downloading image {image_url}: {e}")
    return None

def translate_text(translator, text):
    try:
        translated = translator.translate(text, src='ja', dest='ko')
        return translated.text
    except Exception as e:
        print(f"Translation error: {e}")
        return text

from concurrent.futures import ThreadPoolExecutor

def search_and_crawl(session, brands, categories, output_folder, progress_callback=None, translator=None):
    base_url = "https://www.ecoauc.com/client/auctions/inspect"
    all_items = []
    total_brands = len(brands)

    for brand_index, brand in enumerate(brands):
        params = {
            "limit": "500",
            "sortKey": "1",
            "tableType": "grid",
            "q": brand,
            "low": "",
            "high": "",
            "master_item_brands": "",
            "auction_lane_id": "",
            "master_item_shapes": "",
            "master_item_ranks": "",
            "page": "1"
        }

        for i, category in enumerate(categories):
            params[f"master_item_categories[{i}]"] = category

        total_pages = get_total_pages(session, base_url, params)
        print(f"Total pages for {brand}: {total_pages}")

        image_urls = []
        for page in range(1, total_pages + 1):
            print(f"Crawling page {page} for {brand}...")
            params["page"] = str(page)
            response = session.get(base_url, params=params)
            soup = BeautifulSoup(response.text, 'html.parser')
            
            items = soup.find_all("div", class_="col-sm-6 col-md-4 col-lg-3 mb-grid-card")
            
            if not items:
                print(f"No items found on page {page} for {brand}. Stopping crawl.")
                break

            print(f"Found {len(items)} items on page {page}")

            for idx, item in enumerate(items, 1):
                try:
                    brand_elem = item.find("small", class_="show-case-bland")
                    brand_text = brand_elem.text.strip() if brand_elem else "N/A"
                    if brand_text != brand:
                        continue

                    title_elem = item.find("b")
                    title = title_elem.text.strip() if title_elem else "N/A"
                    title = translate_text(translator, title) if translator else title
                    
                    canopy = item.find("ul", class_="canopy canopy-3 text-default")
                    if canopy:
                        rank_elem = canopy.find_all("li")[0].find("big", class_="canopy-value")
                        rank = rank_elem.next_sibling.strip() if rank_elem else "N/A"
                        
                        price_elem = canopy.find_all("li")[1].find("big", class_="canopy-value")
                        starting_price = price_elem.text.strip() if price_elem else "N/A"
                    else:
                        rank = "N/A"
                        starting_price = "N/A"
                    
                    image_elem = item.find("div", class_="item-image item-image-min pc-image-area")
                    if image_elem and image_elem.find("img"):
                        image_url = image_elem.find("img")['src']
                        image_urls.append(image_url)
                    
                    market_title_elem = item.find("span", class_="market-title")
                    if market_title_elem:
                        market_title_text = market_title_elem.text.strip()
                        market_time = market_title_text
                        market_time = translate_text(translator, market_time) if translator else market_time
                    else:
                        market_time = "N/A"

                    all_items.append({
                        "Brand": brand,
                        "Title": title,
                        "Rank": rank,
                        "Starting Price": starting_price,
                        "Image": None,
                        "Time": market_time
                    })

                    if progress_callback:
                        progress_callback(brand, page, total_pages, len(all_items), idx)

                    if idx % 10 == 0:
                        print(f"Processed {idx}/{len(items)} items on page {page}")
                except Exception as e:
                    print(f"Error processing item {idx} on page {page}: {e}")
                    print(f"Item HTML: {item}")

            print(f"Page {page} crawled successfully.")
            time.sleep(1)

    print(f"Crawling completed. Total items found: {len(all_items)}")

    # 이미지 다운로드
    with ThreadPoolExecutor() as executor:
        futures = []
        for image_url in image_urls:
            futures.append(executor.submit(download_image, image_url, os.path.join(output_folder, 'images')))

        for future in futures:
            image_path = future.result()
            if image_path:
                for item in all_items:
                    if item["Image"] is None:
                        item["Image"] = os.path.basename(image_path)
                        break

    return all_items

def download_image(image_url, save_dir):
    try:
        parsed_url = urllib.parse.urlparse(image_url)
        query_params = urllib.parse.parse_qs(parsed_url.query)
        query_params.pop('w', None)
        query_params.pop('h', None)
        new_query = urllib.parse.urlencode(query_params, doseq=True)
        new_url = urllib.parse.urlunparse(parsed_url._replace(query=new_query))

        response = requests.get(new_url, timeout=30)
        if response.status_code == 200:
            os.makedirs(save_dir, exist_ok=True)
            file_name = os.path.join(save_dir, os.path.basename(new_url))
            with open(file_name, 'wb') as f:
                f.write(response.content)
            return file_name
    except Exception as e:
        print(f"Error downloading image {image_url}: {e}")
    return None

class CrawlerThread(QThread):
    update_progress = pyqtSignal(str, int, int, int, int)
    finished = pyqtSignal(list)

    def __init__(self, session, brands, categories, output_folder, translator):
        QThread.__init__(self)
        self.session = session
        self.brands = brands
        self.categories = categories
        self.output_folder = output_folder
        self.translator = translator

    def run(self):
        items = search_and_crawl(self.session, self.brands, self.categories, self.output_folder, self.progress_callback, self.translator)
        self.finished.emit(items)

    def progress_callback(self, brand, current_page, total_pages, items_found, images_downloaded):
        self.update_progress.emit(brand, current_page, total_pages, items_found, images_downloaded)

class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("브랜드 크롤러")
        self.setGeometry(100, 100, 1200, 800)

        main_widget = QWidget()
        self.setCentralWidget(main_widget)

        layout = QVBoxLayout()
        main_widget.setLayout(layout)

        # 브랜드 선택
        brand_layout = QHBoxLayout()
        self.brand_list = QListWidget()
        self.brand_list.addItems(BRANDS)
        self.brand_list.setSelectionMode(QListWidget.MultiSelection)
        brand_layout.addWidget(self.brand_list)

        # 카테고리 선택
        category_layout = QVBoxLayout()
        category_label = QLabel("카테고리 선택:")
        category_layout.addWidget(category_label)
        self.category_buttons = []
        categories = ["시계", "가방", "귀금속", "악세서리", "소품", "의류", "신발", "기타"]
        category_ids = ["1", "2", "3", "4", "5", "8", "9", "27"]
        for category, category_id in zip(categories, category_ids):
            button = QPushButton(category)
            button.setCheckable(True)
            category_layout.addWidget(button)
            self.category_buttons.append((button, category_id))

        brand_layout.addLayout(category_layout)
        layout.addLayout(brand_layout)

        # 저장 위치 선택
        save_layout = QHBoxLayout()
        self.save_location_label = QLabel("저장 위치: 선택되지 않음")
        save_layout.addWidget(self.save_location_label)
        save_button = QPushButton("저장 위치 선택")
        save_button.clicked.connect(self.select_save_location)
        save_layout.addWidget(save_button)
        layout.addLayout(save_layout)

        # 실행 버튼
        self.run_button = QPushButton("크롤링 시작")
        self.run_button.clicked.connect(self.start_crawling)
        layout.addWidget(self.run_button)

        # 진행 상황 표시
        self.progress_bar = QProgressBar()
        layout.addWidget(self.progress_bar)
        self.progress_label = QLabel("대기 중...")
        layout.addWidget(self.progress_label)

        # 결과 표시 탭
        self.result_tabs = QTabWidget()
        layout.addWidget(self.result_tabs)

        self.output_folder = None
        self.translator = Translator()

    def select_save_location(self):
        folder = QFileDialog.getExistingDirectory(self, "저장 위치 선택")
        if folder:
            self.output_folder = folder
            self.save_location_label.setText(f"저장 위치: {folder}")

    def start_crawling(self):
        if not self.output_folder:
            QMessageBox.warning(self, "경고", "저장 위치를 선택해주세요.")
            return

        selected_brands = [item.text() for item in self.brand_list.selectedItems()]
        if not selected_brands:
            QMessageBox.warning(self, "경고", "브랜드를 선택해주세요.")
            return

        selected_categories = [category_id for button, category_id in self.category_buttons if button.isChecked()]
        if not selected_categories:
            QMessageBox.warning(self, "경고", "카테고리를 선택해주세요.")
            return

        session = login_and_check()
        if not session:
            QMessageBox.critical(self, "오류", "로그인에 실패했습니다.")
            return

        self.crawler_thread = CrawlerThread(session, selected_brands, selected_categories, self.output_folder, self.translator)
        self.crawler_thread.update_progress.connect(self.update_progress)
        self.crawler_thread.finished.connect(self.crawling_finished)
        self.crawler_thread.start()

        self.run_button.setEnabled(False)

    def update_progress(self, brand, current_page, total_pages, items_found, images_downloaded):
        self.progress_label.setText(f"브랜드: {brand}, 페이지: {current_page}/{total_pages}, 아이템: {items_found}, 이미지: {images_downloaded}")
        self.progress_bar.setValue(int(current_page / total_pages * 100))

    def crawling_finished(self, items):
        self.run_button.setEnabled(True)
        self.progress_bar.setValue(100)
        self.progress_label.setText("크롤링 완료")

        # 결과를 표로 표시
        result_table = QTableWidget()
        result_table.setColumnCount(6)
        result_table.setHorizontalHeaderLabels(["Brand", "Title", "Rank", "Starting Price", "Image", "Time"])
        result_table.setRowCount(len(items))

        for row, item in enumerate(items):
            result_table.setItem(row, 0, QTableWidgetItem(item["Brand"]))
            result_table.setItem(row, 1, QTableWidgetItem(item["Title"]))
            result_table.setItem(row, 2, QTableWidgetItem(item["Rank"]))
            result_table.setItem(row, 3, QTableWidgetItem(item["Starting Price"]))
            
            if item["Image"]:
                image_path = os.path.join(self.output_folder, 'images', item["Image"])
                if os.path.exists(image_path):
                    pixmap = QPixmap(image_path)
                    pixmap = pixmap.scaled(QSize(100, 100), Qt.KeepAspectRatio, Qt.SmoothTransformation)
                    label = QLabel()
                    label.setPixmap(pixmap)
                    result_table.setCellWidget(row, 4, label)
                    result_table.setRowHeight(row, 100)
                else:
                    result_table.setItem(row, 4, QTableWidgetItem(item["Image"]))
            else:
                result_table.setItem(row, 4, QTableWidgetItem("No Image"))
            
            result_table.setItem(row, 5, QTableWidgetItem(item["Time"]))

        # 열 너비 조정
        result_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)

        self.result_tabs.addTab(result_table, "크롤링 결과")

        # 엑셀 파일 저장
        self.save_to_excel(items)

        QMessageBox.information(self, "완료", "크롤링이 완료되었고, 결과가 엑셀 파일로 저장되었습니다.")

    def save_to_excel(self, items):
        wb = Workbook()
        ws = wb.active
        headers = ["Brand", "Title", "Rank", "Starting Price", "Image", "Time"]
        ws.append(headers)

        for row, item in enumerate(items, start=2):
            ws.append([item["Brand"], item["Title"], item["Rank"], item["Starting Price"], item["Image"], item["Time"]])
            
            if item["Image"]:
                img_path = os.path.join(self.output_folder, 'images', item["Image"])
                if os.path.exists(img_path):
                    img = Image(img_path)
                    img.width = 100
                    img.height = 100
                    ws.add_image(img, f'E{row}')
                    
                    # Adjust row height to fit the image
                    ws.row_dimensions[row].height = 75

        # Adjust column widths
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 20

        file_path = os.path.join(self.output_folder, "crawling_results.xlsx")
        wb.save(file_path)
        print(f"Data saved to {file_path}")

if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = MainWindow()
    window.show()
    sys.exit(app.exec_())
